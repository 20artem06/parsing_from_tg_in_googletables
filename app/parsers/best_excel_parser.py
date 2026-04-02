from __future__ import annotations

import re
from io import BytesIO

from openpyxl import load_workbook

from app.normalization.aliases import CATEGORY_BY_SHEET
from app.storage.models import RawBestItem
from app.utils.parsing import clean_text, extract_flag, parse_price


HEADER_MODEL_NAMES = {"модель", "model", "наименование"}
HEADER_PRICE_NAMES = {"стоимость", "цена", "price"}
HEADER_FLAG_NAMES = {"флаг", "регион", "flag", "country"}
NOTE_HINTS = {
    "model",
    "price",
    "flag",
    "comment",
    "comments",
    "qty",
    "quantity",
    "stock",
    "total",
}


class BestExcelParser:
    def parse_bytes_by_sheet(self, payload: bytes) -> dict[str, list[RawBestItem]]:
        workbook = load_workbook(filename=BytesIO(payload), data_only=True)
        parsed_by_sheet: dict[str, list[RawBestItem]] = {
            sheet_name: [] for sheet_name in CATEGORY_BY_SHEET
        }

        for sheet_name in workbook.sheetnames:
            if sheet_name not in CATEGORY_BY_SHEET:
                continue

            worksheet = workbook[sheet_name]
            header_row, model_index, price_index, flag_index = self._find_header(worksheet)
            section_path: list[str] = []
            items = parsed_by_sheet[sheet_name]

            for row_number, row in enumerate(
                worksheet.iter_rows(min_row=header_row + 1, values_only=True),
                start=header_row + 1,
            ):
                cells = [clean_text(value) for value in row]
                if not any(cells):
                    section_path = []
                    continue

                model_value = cells[model_index] if model_index < len(cells) else cells[0]
                price_value = row[price_index] if price_index < len(row) else None
                parsed_price = parse_price(price_value)
                flag_value = clean_text(row[flag_index]) if flag_index is not None and flag_index < len(row) else ""

                if parsed_price is None and self._is_sheet_header_row(sheet_name, cells):
                    section_path = []
                    continue

                if parsed_price is None and self._looks_like_section_row(
                    worksheet=worksheet,
                    row_number=row_number,
                    model_index=model_index,
                    price_index=price_index,
                    cells=cells,
                ):
                    section_path = self._update_section_path(section_path, cells)
                    continue

                if not model_value or parsed_price is None:
                    continue

                full_name = self._compose_full_name(section_path, model_value)
                items.append(
                    RawBestItem(
                        sheet_name=sheet_name,
                        row_number=row_number,
                        raw_name=model_value,
                        full_name=full_name,
                        section_path=section_path.copy(),
                        raw_price=clean_text(price_value),
                        price=parsed_price,
                        raw_flag=flag_value or None,
                        country_flag=extract_flag(flag_value) or extract_flag(full_name),
                    )
                )

        return parsed_by_sheet

    def parse_bytes(self, payload: bytes) -> list[RawBestItem]:
        parsed_by_sheet = self.parse_bytes_by_sheet(payload)
        items: list[RawBestItem] = []
        for sheet_name in CATEGORY_BY_SHEET:
            items.extend(parsed_by_sheet.get(sheet_name, []))
        return items

    def _find_header(self, worksheet) -> tuple[int, int, int, int | None]:
        fallback = (1, 0, 1, 2)
        for row_number, row in enumerate(worksheet.iter_rows(min_row=1, max_row=10, values_only=True), start=1):
            cells = [clean_text(value).lower() for value in row]
            if not any(cells):
                continue
            if not any(cell in HEADER_MODEL_NAMES for cell in cells):
                continue
            model_index = next((index for index, cell in enumerate(cells) if cell in HEADER_MODEL_NAMES), 0)
            price_index = next((index for index, cell in enumerate(cells) if cell in HEADER_PRICE_NAMES), 1)
            flag_index = next((index for index, cell in enumerate(cells) if cell in HEADER_FLAG_NAMES), None)
            return row_number, model_index, price_index, flag_index
        return fallback

    def _looks_like_section_row(
        self,
        *,
        worksheet,
        row_number: int,
        model_index: int,
        price_index: int,
        cells: list[str],
    ) -> bool:
        non_empty = [cell for cell in cells if cell]
        if not non_empty:
            return False
        if len(non_empty) > 3:
            return False

        joined = " ".join(non_empty)
        if not any(char.isalnum() for char in joined):
            return False
        if self._looks_like_note_row(joined):
            return False

        for next_row in worksheet.iter_rows(
            min_row=row_number + 1,
            max_row=min(row_number + 3, worksheet.max_row),
            values_only=True,
        ):
            next_cells = [clean_text(value) for value in next_row]
            if not any(next_cells):
                continue
            next_model = next_cells[model_index] if model_index < len(next_cells) else next_cells[0]
            next_price_value = next_row[price_index] if price_index < len(next_row) else None
            if next_model and parse_price(next_price_value) is not None:
                return True

        return False

    def _looks_like_note_row(self, text: str) -> bool:
        normalized = clean_text(text).lower()
        if not normalized:
            return False
        if normalized.endswith(":"):
            return True
        return re.search(r"\b(?:%s)\b" % "|".join(re.escape(hint) for hint in NOTE_HINTS), normalized) is not None

    def _is_sheet_header_row(self, sheet_name: str, cells: list[str]) -> bool:
        non_empty = [cell for cell in cells if cell]
        if len(non_empty) != 1:
            return False
        return clean_text(non_empty[0]).lower() == clean_text(sheet_name).lower()

    def _update_section_path(self, current: list[str], cells: list[str]) -> list[str]:
        parts = [cell for cell in cells if cell]
        if not parts:
            return current
        joined = " ".join(parts).lower()
        if any(keyword in joined for keyword in ("iphone", "ipad", "macbook", "airpods", "watch", "magic keyboard")):
            return parts
        updated = [value for value in current if value.lower() not in joined]
        for part in parts:
            if part not in updated:
                updated.append(part)
        return updated[-3:]

    def _compose_full_name(self, section_path: list[str], raw_name: str) -> str:
        parts = []
        raw_lower = raw_name.lower()
        for section in section_path:
            if section.lower() not in raw_lower:
                parts.append(section)
        parts.append(raw_name)
        return " ".join(part for part in parts if part).strip()
