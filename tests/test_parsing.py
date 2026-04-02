from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook

from app.parsers.best_excel_parser import BestExcelParser
from app.parsers.sonic_text_parser import SonicTextParser
from app.storage.models import SonicBatch
from app.utils.parsing import parse_price


US_FLAG = "\U0001F1FA\U0001F1F8"


def test_parse_price_variants() -> None:
    assert parse_price("27.900") == 27900
    assert parse_price("43 500") == 43500
    assert parse_price("52,900") == 52900
    assert parse_price(61000) == 61000


def test_parse_sonic_line_and_model_code() -> None:
    parser = SonicTextParser()
    item = parser.parse_line(
        line=f"Magic Keyboard Air 11 Black (MGYX4) - 25.500 {US_FLAG}",
        line_number=3,
        current_section="Magic Keyboard (iPad Air)",
        batch_message_ids=[101, 102],
    )

    assert item is not None
    assert item.price == 25500
    assert item.country_flag == US_FLAG
    assert item.model_code == "MGYX4"
    assert item.full_name.startswith("Magic Keyboard (iPad Air)")


def test_parse_best_excel_with_sections() -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "iPad"
    sheet.append(["Модель", "Стоимость", "Флаг"])
    sheet.append(["iPad 11 (A16) 2025", None, None])
    sheet.append(["128GB", None, None])
    sheet.append(["Silver Wi-Fi", "27.900", US_FLAG])

    buffer = BytesIO()
    workbook.save(buffer)

    parser = BestExcelParser()
    items = parser.parse_bytes(buffer.getvalue())

    assert len(items) == 1
    assert items[0].price == 27900
    assert items[0].country_flag == US_FLAG
    assert items[0].full_name == "iPad 11 (A16) 2025 128GB Silver Wi-Fi"


def test_parse_best_excel_resets_section_after_blank_separator() -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Аксессуары Apple"
    sheet.append(["Модель", "Стоимость", "Флаг"])
    sheet.append(["Аксессуары Apple", None, None])
    sheet.append(["64GB", None, None])
    sheet.append(["Apple TV 64GB", "13.800", None])
    sheet.append([None, None, None])
    sheet.append(["128GB", None, None])
    sheet.append(["Apple TV 4K 128GB", "15.200", None])
    sheet.append([None, None, None])
    sheet.append(["Pencil USB-C", "7.500", None])

    buffer = BytesIO()
    workbook.save(buffer)

    parser = BestExcelParser()
    items = parser.parse_bytes(buffer.getvalue())

    assert len(items) == 3
    assert items[0].full_name == "Apple TV 64GB"
    assert items[1].full_name == "Apple TV 4K 128GB"
    assert items[2].full_name == "Pencil USB-C"


def test_parse_best_excel_ignores_note_like_row_as_section() -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "iPad"
    sheet.append(["Модель", "Стоимость", "Флаг"])
    sheet.append(["iPad 11 (A16) 2025", None, None])
    sheet.append(["Comment:", None, None])
    sheet.append(["128GB Silver Wi-Fi", "27.900", US_FLAG])

    buffer = BytesIO()
    workbook.save(buffer)

    parser = BestExcelParser()
    items = parser.parse_bytes(buffer.getvalue())

    assert len(items) == 1
    assert items[0].full_name == "iPad 11 (A16) 2025 128GB Silver Wi-Fi"


def test_parse_sonic_batch_keeps_section_context() -> None:
    batch = SonicBatch(
        message_ids=[201, 202],
        raw_text=(
            "iPad 11 (A16) 2025\n"
            f"iPad 11 128 Silver Wi-Fi - 27.900 {US_FLAG}\n"
            f"iPad 11 128 Pink LTE - 43.500 {US_FLAG}"
        ),
    )
    parser = SonicTextParser()
    items = parser.parse_batch(batch)

    assert len(items) == 2
    assert items[0].section_name == "iPad 11 (A16) 2025"
    assert items[1].price == 43500


def test_single_line_sonic_message_is_parsed() -> None:
    batch = SonicBatch(
        message_ids=[301],
        raw_text=f"iPad 11 256 Pink Wi-Fi - 35.900 {US_FLAG}",
    )
    parser = SonicTextParser()
    items = parser.parse_batch(batch)

    assert len(items) == 1
    assert items[0].full_name == "iPad 11 256 Pink Wi-Fi"
    assert items[0].price == 35900


def test_separator_only_sonic_line_does_not_become_section() -> None:
    batch = SonicBatch(
        message_ids=[401],
        raw_text=(
            "➖➖➖➖ ➖➖➖➖➖➖➖➖➖➖\n\n"
            "AirTag 1pc - 2600\n"
            "AirTag 4pc - 7500\n"
        ),
    )
    parser = SonicTextParser()
    items = parser.parse_batch(batch)

    assert len(items) == 2
    assert items[0].section_name is None
    assert items[0].full_name == "AirTag 1pc"
    assert items[1].full_name == "AirTag 4pc"


def test_placeholder_only_sonic_message_supports_more_than_dots() -> None:
    parser = SonicTextParser()

    assert parser._is_placeholder_message("•\n•") is True
    assert parser._is_placeholder_message("---\n———") is True
    assert parser._is_placeholder_message("закрыто") is True
    assert parser._is_placeholder_message("iPad 11 128 Silver Wi-Fi - 27.900") is False
